from __future__ import annotations

import csv
import io
import json
import sqlite3
from dataclasses import dataclass, field
from typing import Iterable

from openpyxl import Workbook, load_workbook

CSV_HEADER = [
    "question_number",
    "prompt",
    "type",
    "choice_a",
    "choice_b",
    "choice_c",
    "choice_d",
    "choice_e",
    "correct",
    "points",
]

CHOICE_KEYS = ["choice_a", "choice_b", "choice_c", "choice_d", "choice_e"]
VALID_TYPES = {"mcq", "true_false", "short"}


@dataclass
class ImportResult:
    imported: int = 0
    errors: list[str] = field(default_factory=list)


def _normalize_type(raw: str | None) -> str:
    value = (raw or "mcq").strip().lower().replace("-", "_").replace(" ", "_")
    aliases = {"mc": "mcq", "multiple_choice": "mcq", "tf": "true_false", "truefalse": "true_false"}
    return aliases.get(value, value)


def _normalize_correct(qtype: str, raw: str | None) -> str | None:
    if raw is None:
        return None
    value = str(raw).strip()
    if not value:
        return None
    if qtype == "mcq":
        letter = value[0].upper()
        if letter not in "ABCDE":
            raise ValueError(f"correct must be A–E for mcq, got {value!r}")
        return letter
    if qtype == "true_false":
        lowered = value.lower()
        if lowered in {"true", "t", "yes", "1"}:
            return "true"
        if lowered in {"false", "f", "no", "0"}:
            return "false"
        raise ValueError(f"correct must be true/false for true_false, got {value!r}")
    return value


def _choices_from_row(row: dict, qtype: str) -> list[str]:
    if qtype == "true_false":
        return ["True", "False"]
    if qtype == "short":
        return []
    choices = []
    for key in CHOICE_KEYS:
        text = str(row.get(key) or "").strip()
        if text:
            choices.append(text)
    return choices


def _row_to_question(row: dict, index: int) -> dict:
    prompt = str(row.get("prompt") or "").strip()
    if not prompt:
        raise ValueError("prompt is required")
    qtype = _normalize_type(row.get("type"))
    if qtype not in VALID_TYPES:
        raise ValueError(f"unknown type {qtype!r}")
    choices = _choices_from_row(row, qtype)
    if qtype == "mcq" and len(choices) < 2:
        raise ValueError("mcq needs at least two choices")
    correct = _normalize_correct(qtype, row.get("correct"))
    if qtype == "mcq" and correct:
        idx = ord(correct) - ord("A")
        if idx >= len(choices):
            raise ValueError(f"correct {correct} has no matching choice")
    points_raw = row.get("points")
    points = float(points_raw) if str(points_raw or "").strip() else 1.0
    position_raw = row.get("question_number")
    position = int(position_raw) if str(position_raw or "").strip() else index
    return {
        "position": position,
        "prompt": prompt,
        "type": qtype,
        "choices": choices,
        "correct": correct,
        "points": points,
    }


def parse_csv(text: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV is missing a header row")
    return [dict(row) for row in reader]


def parse_xlsx(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration as exc:
        raise ValueError("Excel sheet is empty") from exc
    rows = []
    for values in rows_iter:
        row = {header[i]: (values[i] if i < len(values) else None) for i in range(len(header))}
        if not any(v not in (None, "") for v in row.values()):
            continue
        rows.append(row)
    return rows


def import_table(conn: sqlite3.Connection, set_id: int, rows: Iterable[dict]) -> ImportResult:
    result = ImportResult()
    taken = {
        row[0]
        for row in conn.execute("SELECT position FROM questions WHERE set_id = ?", (set_id,))
    }
    next_pos = (max(taken) + 1) if taken else 1
    for i, raw in enumerate(rows, start=1):
        try:
            q = _row_to_question(raw, next_pos)
            pos = q["position"]
            if pos in taken:
                pos = next_pos
            conn.execute(
                """
                INSERT INTO questions (set_id, position, prompt, type, choices_json, correct, points, visible)
                VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                """,
                (
                    set_id,
                    pos,
                    q["prompt"],
                    q["type"],
                    json.dumps(q["choices"]),
                    q["correct"],
                    q["points"],
                ),
            )
            taken.add(pos)
            next_pos = max(next_pos, pos) + 1
            result.imported += 1
        except Exception as exc:  # noqa: BLE001 — collect row errors for the teacher
            result.errors.append(f"Row {i}: {exc}")
    conn.commit()
    return result


def export_rows(conn: sqlite3.Connection, set_id: int) -> list[dict]:
    questions = conn.execute(
        "SELECT * FROM questions WHERE set_id = ? ORDER BY position, id",
        (set_id,),
    ).fetchall()
    rows = []
    for q in questions:
        choices = json.loads(q["choices_json"] or "[]")
        padded = list(choices) + [""] * (5 - len(choices))
        rows.append(
            {
                "question_number": q["position"],
                "prompt": q["prompt"],
                "type": q["type"],
                "choice_a": padded[0] if q["type"] == "mcq" else "",
                "choice_b": padded[1] if q["type"] == "mcq" else "",
                "choice_c": padded[2] if q["type"] == "mcq" else "",
                "choice_d": padded[3] if q["type"] == "mcq" else "",
                "choice_e": padded[4] if q["type"] == "mcq" else "",
                "correct": q["correct"] or "",
                "points": q["points"],
            }
        )
    return rows


def export_csv(conn: sqlite3.Connection, set_id: int) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_HEADER, lineterminator="\n")
    writer.writeheader()
    writer.writerows(export_rows(conn, set_id))
    return buf.getvalue()


def export_xlsx(conn: sqlite3.Connection, set_id: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "questions"
    ws.append(CSV_HEADER)
    for row in export_rows(conn, set_id):
        ws.append([row[h] for h in CSV_HEADER])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
